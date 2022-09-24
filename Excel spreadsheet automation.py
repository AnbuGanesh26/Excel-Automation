import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row+1):
        old_cell = sheet.cell(row, 3)
        corrected_price_cell = sheet.cell(row, 4)
        revised_price_cell = sheet.cell(row, 5)
        corrected_price = old_cell.value * 0.9
        corrected_price_cell.value = corrected_price
        revised_price = old_cell.value * 2
        revised_price_cell.value = revised_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=3,
                       max_col=5)

    sheet['D1'] = "corrected price"
    sheet['E1'] = "New price"

    newchart = BarChart()
    newchart.add_data(values)
    sheet.add_chart(newchart, 'c6')

    wb.save(filename)

process_workbook('transactions.xlsx')

